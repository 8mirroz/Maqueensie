#!/usr/bin/env python3
"""Materialize tier rule markdown guides from canonical XLSX source."""

from __future__ import annotations

import argparse
import datetime as dt
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: 'openpyxl' module not found. Please install it using: pip install openpyxl")
    sys.exit(1)

from pathlib import Path
from typing import Dict, List

ROOT = Path(__file__).resolve().parents[1]
XLSX_PATH = ROOT / "obsidian_tier_rules_spec.xlsx"
OUT_DIR = ROOT / "10_Гайды_и_Правила"
DEBUG = False

def log(msg: str):
    if DEBUG:
        print(f"[DEBUG] {msg}")

SHEET_CONFIG = {
    "tier_rules_master": {
        "output": "tier_rules.md",
        "title": "tier_rules_master",
        "prefix": "TR-",
        "doc_id": "tier_rules_master",
        "extra_frontmatter": {
            "doc_type": "canonical_rules_table",
            "version": "1.0.0",
            "currency": "RUB",
            "geo_scope_default": "[Moscow, Moscow_Oblast]",
            "metrics_allowed": "[rub_m2, package_rub]",
            "status": "active",
        },
    },
    "tier_rules_exceptions": {
        "output": "tier_rules_exceptions.md",
        "title": "tier_rules_exceptions",
        "prefix": "EXC-",
        "doc_id": "tier_rules_exceptions",
        "extra_frontmatter": {"version": "1.0.0", "status": "active"},
    },
    "tier_rules_changelog": {
        "output": "tier_rules_changelog.md",
        "title": "tier_rules_changelog",
        "prefix": "CHG-",
        "doc_id": "tier_rules_changelog",
        "extra_frontmatter": {"version": "1.0.0", "status": "active"},
    },
    "tier_rules_evidence": {
        "output": "tier_rules_evidence.md",
        "title": "tier_rules_evidence",
        "prefix": "EVD-",
        "doc_id": "tier_rules_evidence",
        "extra_frontmatter": {"version": "1.0.0", "status": "active"},
    },
}

LIST_FIELDS = {
    "tier_rules_master": {"source_urls", "source_access_dates", "snapshot_hashes"},
    "tier_rules_exceptions": {"source_urls", "source_access_dates", "snapshot_hashes"},
    "tier_rules_changelog": set(),
    "tier_rules_evidence": set(),
}

DATE_FIELDS = {
    "tier_rules_master": {"effective_from", "effective_to"},
    "tier_rules_exceptions": {"date_added", "date_verified"},
    "tier_rules_changelog": set(),
    "tier_rules_evidence": {"access_date"},
}

STATUS_WHITELIST = {"draft", "active", "draft_conflict", "deprecated"}


def _normalize_for_check(text: str) -> str:
    """Ignore volatile fields that change between runs."""
    out = []
    for line in text.splitlines():
        if line.startswith("last_materialized:"):
            continue
        out.append(line)
    return "\n".join(out).strip() + "\n"


def _sheet_rows(xlsx_path: Path, sheet_name: str) -> List[List[str]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    sheet = wb[sheet_name]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append([str(cell if cell is not None else "") for cell in row])
    return rows


def _normalize_date(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    for fmt in ("%Y-%m-%d", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%dT%H:%M:%S%z"):
        try:
            return dt.datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    if "T" in value:
        return value.split("T", 1)[0]
    return value


def _normalize_cell(sheet: str, column: str, value: str) -> str:
    value = value.strip()
    if column in LIST_FIELDS[sheet]:
        parts = [p.strip() for p in value.replace("\r", "").split("\n") if p.strip()]
        if column == "source_access_dates":
            parts = [_normalize_date(p) for p in parts]
        return "; ".join(parts)

    if column in DATE_FIELDS[sheet]:
        return _normalize_date(value)

    if sheet == "tier_rules_master" and column == "status":
        status = value.lower()
        return status if status in STATUS_WHITELIST else "draft"

    if sheet == "tier_rules_exceptions" and column == "status":
        status = value.lower()
        return status if status in STATUS_WHITELIST else "draft"

    return value


def _render_markdown_table(headers: List[str], rows: List[List[str]]) -> str:
    escaped_headers = [h.replace("|", "\\|") for h in headers]
    out = [
        "| " + " | ".join(escaped_headers) + " |",
        "|" + "|".join(["---"] * len(headers)) + "|",
    ]
    for row in rows:
        padded = row + [""] * (len(headers) - len(row))
        cells = [str(cell).replace("|", "\\|") for cell in padded[: len(headers)]]
        out.append("| " + " | ".join(cells) + " |")
    return "\n".join(out)


def _trim_header(header: List[str]) -> List[str]:
    last_non_empty = -1
    for idx, value in enumerate(header):
        if value.strip():
            last_non_empty = idx
    if last_non_empty == -1:
        return []
    return [h.strip() for h in header[: last_non_empty + 1]]


def _render_doc(sheet: str) -> str:
    cfg = SHEET_CONFIG[sheet]
    rows = _sheet_rows(XLSX_PATH, sheet)
    if not rows:
        raise RuntimeError(f"No rows in sheet: {sheet}")

    header = _trim_header(rows[0])
    data = []
    for row in rows[1:]:
        if not row:
            continue
        first = (row[0] if len(row) > 0 else "").strip()
        if not first.startswith(cfg["prefix"]):
            continue
        norm = []
        padded = row[: len(header)] + [""] * max(0, len(header) - len(row))
        for key, raw in zip(header, padded[: len(header)]):
            norm.append(_normalize_cell(sheet, key, str(raw)))
        data.append(norm)

    fm = [
        "---",
        f"doc_id: {cfg['doc_id']}",
    ]
    for key, value in cfg["extra_frontmatter"].items():
        fm.append(f"{key}: {value}")
    fm.extend([
        f"source_file: {XLSX_PATH.name}",
        f"last_materialized: {dt.date.today().isoformat()}",
        "---",
        "",
        f"# {cfg['title']}",
        "",
    ])

    if sheet == "tier_rules_master":
        fm.extend(
            [
                "## Validation Logic",
                "",
                "- `economy_max_rub < comfort_min_rub <= comfort_max_rub < business_min_rub <= business_max_rub < premium_min_rub`",
                "- Tier ranges must not overlap",
                "- `status=active` only if `source_count_verified >= 2`",
                "- `metric_type=package_rub` requires empty `area_band_m2_min/max`",
                "- `effective_to` must be empty or `>= effective_from`",
                "- Any threshold update creates a new `rule_id`",
                "",
            ]
        )

    table = _render_markdown_table(header, data)

    qa = [
        "",
        "## QA Checks",
        "",
        "- [x] Canonical source is `obsidian_tier_rules_spec.xlsx`",
        "- [x] Multiline list fields normalized with `;` separator",
        "- [x] Date fields normalized to `YYYY-MM-DD` where applicable",
        "- [x] Status values constrained by whitelist",
        "",
    ]

    return "\n".join(fm) + table + "\n" + "\n".join(qa)


def materialize(check: bool) -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    mismatches = []

    if not XLSX_PATH.exists():
        print(f"ERROR: Specification file not found at {XLSX_PATH}")
        return 1

    for sheet, cfg in SHEET_CONFIG.items():
        try:
            rendered = _render_doc(sheet)
            path = OUT_DIR / cfg["output"]
            if check:
                current = path.read_text(encoding="utf-8") if path.exists() else ""
                if _normalize_for_check(current) != _normalize_for_check(rendered):
                    mismatches.append(str(path))
                continue
            
            path.write_text(rendered, encoding="utf-8")
            print(f"SUCCESS: Materialized {cfg['output']}")
        except Exception as e:
            print(f"ERROR: Failed to process sheet '{sheet}': {e}")
            if DEBUG:
                import traceback
                traceback.print_exc()
            return 1

    if check:
        if mismatches:
            print("Materialization drift detected:")
            for item in mismatches:
                print(f"- {item}")
            return 1
        else:
            print("SUCCESS: No materialization drift detected.")
    
    return 0


def main() -> int:
    global DEBUG
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--check", action="store_true", help="Validate generated files without writing")
    parser.add_argument("--debug", action="store_true", help="Enable verbose debug logging")
    args = parser.parse_args()
    DEBUG = args.debug
    return materialize(check=args.check)


if __name__ == "__main__":
    raise SystemExit(main())
