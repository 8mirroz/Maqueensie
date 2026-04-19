import openpyxl
from pathlib import Path
import re

xlsx_path = Path("obsidian_tier_rules_spec.xlsx")
wb = openpyxl.load_workbook(xlsx_path)
sheet = wb["tier_rules_master"]

def get_col_idx(header_name):
    headers = [cell.value for cell in sheet[1]]
    return headers.index(header_name) + 1

cols = {
    "rule_id": get_col_idx("rule_id"),
    "domain": get_col_idx("domain"),
    "subdomain": get_col_idx("subdomain_scope"),
    "economy_min": get_col_idx("economy_min_rub"),
    "economy_max": get_col_idx("economy_max_rub"),
    "comfort_min": get_col_idx("comfort_min_rub"),
    "comfort_max": get_col_idx("comfort_max_rub"),
    "business_min": get_col_idx("business_min_rub"),
    "business_max": get_col_idx("business_max_rub"),
    "premium_min": get_col_idx("premium_min_rub"),
    "premium_max": get_col_idx("premium_max_rub"),
}

# 1. Update Renovation (TR-REN)
# Values from 04_Repair_Price_Matrix.md:
# Economy: 12-20, Comfort: 20-35, Business: 35-65, Premium: 70-150
for row in range(2, sheet.max_row + 1):
    rule_id = sheet.cell(row=row, column=cols["rule_id"]).value
    domain = sheet.cell(row=row, column=cols["domain"]).value
    
    if domain == "Renovation" or (rule_id and rule_id.startswith("TR-REN")):
        sheet.cell(row=row, column=cols["economy_min"]).value = 12000
        sheet.cell(row=row, column=cols["economy_max"]).value = 20000
        sheet.cell(row=row, column=cols["comfort_min"]).value = 20001
        sheet.cell(row=row, column=cols["comfort_max"]).value = 35000
        sheet.cell(row=row, column=cols["business_min"]).value = 35001
        sheet.cell(row=row, column=cols["business_max"]).value = 65000
        sheet.cell(row=row, column=cols["premium_min"]).value = 70000
        sheet.cell(row=row, column=cols["premium_max"]).value = 150000
        print(f"Updated Renovation rule: {rule_id}")

# 2. Update Interior Design (TR-INT)
# Refined values based on residential-interiors__moscow-mo__pricing_table__apartments_houses.md
# Economy: 1000-3000, Comfort: 3000-5000, Business: 5000-8000, Premium: 8000-15000
for row in range(2, sheet.max_row + 1):
    rule_id = sheet.cell(row=row, column=cols["rule_id"]).value
    domain = sheet.cell(row=row, column=cols["domain"]).value
    
    if domain == "Interior Design" or (rule_id and rule_id.startswith("TR-INT")):
        sheet.cell(row=row, column=cols["economy_min"]).value = 1000
        sheet.cell(row=row, column=cols["economy_max"]).value = 3000
        sheet.cell(row=row, column=cols["comfort_min"]).value = 3001
        sheet.cell(row=row, column=cols["comfort_max"]).value = 5000
        sheet.cell(row=row, column=cols["business_min"]).value = 5001
        sheet.cell(row=row, column=cols["business_max"]).value = 8000
        sheet.cell(row=row, column=cols["premium_min"]).value = 8001
        sheet.cell(row=row, column=cols["premium_max"]).value = 15000
        print(f"Updated Interior Design rule: {rule_id}")

wb.save(xlsx_path)
print("Synchronization completed.")
