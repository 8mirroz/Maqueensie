import openpyxl
from pathlib import Path

xlsx_path = Path("obsidian_tier_rules_spec.xlsx")
wb = openpyxl.load_workbook(xlsx_path)

def update_sheet_schema(sheet):
    headers = [cell.value for cell in sheet[1]]
    if 'budget_min_rub' not in headers:
        return
    
    # Column indices for budget_max_rub (1-indexed)
    budget_max_idx = headers.index('budget_max_rub') + 1
    
    # Rename budget to economy
    sheet.cell(row=1, column=headers.index('budget_min_rub') + 1).value = 'economy_min_rub'
    sheet.cell(row=1, column=headers.index('budget_max_rub') + 1).value = 'economy_max_rub'
    
    # Insert two columns after economy_max_rub
    sheet.insert_cols(budget_max_idx + 1, 2)
    sheet.cell(row=1, column=budget_max_idx + 1).value = 'comfort_min_rub'
    sheet.cell(row=1, column=budget_max_idx + 2).value = 'comfort_max_rub'
    
    # Initialize some values if needed, otherwise leave blank or 0
    # For now, we'll just leave them for the sync script to fill.
    print(f"Updated schema for sheet: {sheet.title}")

update_sheet_schema(wb["tier_rules_master"])
if "tier_rules_exceptions" in wb.sheetnames:
    update_sheet_schema(wb["tier_rules_exceptions"])

wb.save(xlsx_path)
print("Consolidated schema update completed.")
